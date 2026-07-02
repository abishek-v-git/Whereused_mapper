import io
import re
import openpyxl
import pandas as pd
from django.shortcuts import render, redirect
from django.http import HttpResponse


WHEREUSED_FILTER_PHASE = "Production"
WHEREUSED_FILTER_TYPES = {"Assembly", "Top Assembly"}
WHEREUSED_DISPLAY_COLS = ["Parent Item", "Project Name", "Product Line(s)"]

_QUALIFIER_RE = re.compile(r'\s*\([^)]*\)\s*$')


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _normalize_df(df):
    df.columns = df.columns.str.strip()
    return df.apply(lambda col: col.str.strip() if col.dtype == object else col)


def _find_header_row(file_obj, target_col):
    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(str(cell).strip() == target_col for cell in row if cell is not None):
            wb.close()
            return i
    wb.close()
    return 0


def _read_requirements(file_obj):
    header_row = _find_header_row(file_obj, "PCA / SKU")
    file_obj.seek(0)
    df = pd.read_excel(file_obj, header=header_row, dtype=str)
    return _normalize_df(df)


def _build_pca_model_map(req_df):
    """Returns {PCA_UPPER: {model_lower, ...}} from requirements."""
    pca_to_models = {}
    for _, row in req_df.iterrows():
        pca_raw = str(row.get("PCA / SKU") or "")
        model_raw = str(row.get("MODEL") or "")

        models = set()
        for part in model_raw.split("/"):
            cleaned = _QUALIFIER_RE.sub("", part.strip()).strip().lower()
            if cleaned:
                models.add(cleaned)

        for pca_part in pca_raw.split("/"):
            pca_stem = pca_part.strip().upper()
            if pca_stem:
                pca_to_models.setdefault(pca_stem, set()).update(models)

    return pca_to_models


def _match_project_name(project_name, models_set):
    """
    Returns True if project_name matches any model in models_set.
    Checks both directions because the relationship can go either way:
      - Model is a prefix of the project name: "OTTER" → "Otter Lake-S"
      - Project name is a prefix of the model:  "Gardena" ← "GARDENAE"
    """
    pn = str(project_name).lower().strip() if pd.notna(project_name) else ""
    if not pn:
        return False
    for m in models_set:
        if pn == m or pn.startswith(m) or m.startswith(pn):
            return True
    return False


# ---------------------------------------------------------------------------
# Format 1 — Flat single-PCA files
#   Columns: Parent Item | Lifecycle Phase | Project Name F | Product Line(s) F
#   PCA identified by filename stem
# ---------------------------------------------------------------------------

def _read_whereused_flat(file_obj):
    header_row = _find_header_row(file_obj, "Parent Item")
    file_obj.seek(0)
    df = pd.read_excel(file_obj, header=header_row, dtype=str)
    return _normalize_df(df)


def _filter_flat_df(df, models_for_pca):
    filtered = df[
        (df["Lifecycle Phase"] == WHEREUSED_FILTER_PHASE) &
        df["Project Name F"].apply(lambda pn: _match_project_name(pn, models_for_pca))
    ][["Parent Item", "Project Name F", "Product Line(s) F"]].copy()
    filtered.columns = WHEREUSED_DISPLAY_COLS
    filtered.reset_index(drop=True, inplace=True)
    return filtered


def _flat_result(wu_file, filtered):
    return {
        "filename": wu_file.name,
        "pca": wu_file.name.rsplit(".", 1)[0],
        "format": "flat",
        "rows": [
            [str(v) if pd.notna(v) else "" for v in row]
            for row in filtered.values.tolist()
        ],
        "row_count": len(filtered),
    }


def _process_flat_files(flat_files, pca_to_models):
    results, skipped = [], []
    pca_files = [f for f in flat_files if not f.name.rsplit(".", 1)[0].strip().upper().startswith("ASY")]
    asy_files = [f for f in flat_files if f.name.rsplit(".", 1)[0].strip().upper().startswith("ASY")]

    parent_item_projects = {}  # {PARENT_ITEM_UPPER: cleaned_project_name}

    for wu_file in pca_files:
        stem = wu_file.name.rsplit(".", 1)[0].strip().upper()

        if stem not in pca_to_models:
            skipped.append(wu_file.name)
            continue

        models_for_pca = pca_to_models[stem]
        df = _read_whereused_flat(wu_file)

        required = {"Parent Item", "Lifecycle Phase", "Project Name F", "Product Line(s) F"}
        missing = required - set(df.columns)
        if missing:
            raise ValueError(
                f"Flat file '{wu_file.name}' is missing columns: {', '.join(sorted(missing))}"
            )

        production = df[df["Lifecycle Phase"] == WHEREUSED_FILTER_PHASE]
        for _, row in production.iterrows():
            parent_item = str(row.get("Parent Item") or "").strip().upper()
            project_name = row.get("Project Name F")
            if parent_item and pd.notna(project_name):
                cleaned = _QUALIFIER_RE.sub("", str(project_name).strip()).strip().lower()
                if cleaned:
                    parent_item_projects[parent_item] = cleaned

        filtered = _filter_flat_df(df, models_for_pca)
        results.append(_flat_result(wu_file, filtered))

    for wu_file in asy_files:
        stem = wu_file.name.rsplit(".", 1)[0].strip().upper()

        if stem not in parent_item_projects:
            skipped.append(f"{wu_file.name} (parent item not found in a Production row of the uploaded PCA files)")
            continue

        models_for_pca = {parent_item_projects[stem]}
        df = _read_whereused_flat(wu_file)

        required = {"Parent Item", "Lifecycle Phase", "Project Name F", "Product Line(s) F"}
        missing = required - set(df.columns)
        if missing:
            raise ValueError(
                f"Flat file '{wu_file.name}' is missing columns: {', '.join(sorted(missing))}"
            )

        filtered = _filter_flat_df(df, models_for_pca)
        results.append(_flat_result(wu_file, filtered))

    return results, skipped


# ---------------------------------------------------------------------------
# Format 2 — Where Used Report (multi-section)
#   Each section: Item Number: PCA-XXXXX / Description: / column headers / data
#   Columns: Level | Part Type / Document Type | Number | Lifecycle Phase |
#            Project Name | Product Line(s) | ...
# ---------------------------------------------------------------------------

def _parse_whereused_report(file_obj):
    """
    Parses a Where Used Report Excel that may contain multiple PCA sections.
    Returns list of {pca: str, df: DataFrame}.
    """
    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    def cs(c):
        return str(c).strip() if c is not None else ""

    sections = []
    i = 0
    while i < len(rows):
        row_cells = [cs(c) for c in rows[i]]
        first = row_cells[0] if row_cells else ""

        if first == "Item Number:":
            pca = row_cells[1] if len(row_cells) > 1 else ""
            i += 1
            if i < len(rows):
                i += 1  # skip Description row

            # Find column-header row
            header = None
            while i < len(rows):
                r = [cs(c) for c in rows[i]]
                if "Level" in r and "Part Type / Document Type" in r:
                    header = r
                    i += 1
                    break
                i += 1

            if header is None:
                continue

            data = []
            while i < len(rows):
                r = [cs(c) for c in rows[i]]
                if r[0] == "Item Number:":
                    break
                if any(r):
                    data.append(r)
                i += 1

            if pca and data:
                ncols = len(header)
                padded = [r[:ncols] + [""] * max(0, ncols - len(r)) for r in data]
                df = pd.DataFrame(padded, columns=header)
                df = df.apply(lambda col: col.str.strip() if col.dtype == object else col)
                sections.append({"pca": pca, "df": df})
        else:
            i += 1

    return sections


def _process_report_files(report_files, pca_to_models):
    results, skipped = [], []

    for wu_file in report_files:
        sections = _parse_whereused_report(wu_file)

        if not sections:
            skipped.append(f"{wu_file.name} (no parseable sections found)")
            continue

        for section in sections:
            pca_key = section["pca"].strip().upper()

            if pca_key not in pca_to_models:
                skipped.append(section["pca"])
                continue

            models_for_pca = pca_to_models[pca_key]
            df = section["df"]

            required = {
                "Part Type / Document Type", "Lifecycle Phase",
                "Number", "Project Name", "Product Line(s)"
            }
            missing = required - set(df.columns)
            if missing:
                raise ValueError(
                    f"Section '{section['pca']}' in '{wu_file.name}' is missing "
                    f"columns: {', '.join(sorted(missing))}"
                )

            filtered = df[
                df["Part Type / Document Type"].isin(WHEREUSED_FILTER_TYPES) &
                (df["Lifecycle Phase"] == WHEREUSED_FILTER_PHASE) &
                df["Project Name"].apply(lambda pn: _match_project_name(pn, models_for_pca))
            ][["Number", "Project Name", "Product Line(s)"]].copy()
            filtered.columns = WHEREUSED_DISPLAY_COLS
            filtered.reset_index(drop=True, inplace=True)

            results.append({
                "filename": wu_file.name,
                "pca": section["pca"],
                "format": "report",
                "rows": [
                    [str(v) if pd.notna(v) else "" for v in row]
                    for row in filtered.values.tolist()
                ],
                "row_count": len(filtered),
            })

    return results, skipped


# ---------------------------------------------------------------------------
# Main analysis entry point
# ---------------------------------------------------------------------------

def _run_analysis(req_file, flat_files, report_files):
    req_df = _read_requirements(req_file)

    for col in ("PCA / SKU", "MODEL"):
        if col not in req_df.columns:
            raise ValueError(f"Requirements file is missing the '{col}' column.")

    pca_to_models = _build_pca_model_map(req_df)

    results, skipped = [], []

    if flat_files:
        r, s = _process_flat_files(flat_files, pca_to_models)
        results.extend(r)
        skipped.extend(s)

    if report_files:
        r, s = _process_report_files(report_files, pca_to_models)
        results.extend(r)
        skipped.extend(s)

    return results, skipped


# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------

def upload(request):
    if request.method == "POST":
        req_file  = request.FILES.get("requirements_file")
        flat_files   = request.FILES.getlist("whereused_files")
        report_files = request.FILES.getlist("whereused_report_files")

        flash = {}

        if not req_file:
            flash["error"] = "Please upload the requirements file."
        elif not flat_files and not report_files:
            flash["error"] = "Please upload at least one Where-Used file (flat or report)."
        else:
            try:
                results, skipped = _run_analysis(req_file, flat_files, report_files)

                if skipped:
                    flash["warning"] = (
                        f"{len(skipped)} PCA(s) had no matching entry in requirements and were skipped: "
                        + ", ".join(skipped)
                    )

                if results:
                    total_rows = sum(r["row_count"] for r in results)
                    all_rows = []
                    for r in results:
                        for row in r["rows"]:
                            all_rows.append([r["pca"]] + row)

                    request.session["result_data"] = {
                        "columns": ["PCA / SKU"] + WHEREUSED_DISPLAY_COLS,
                        "rows": all_rows,
                    }
                    flash.update({
                        "results": results,
                        "columns": WHEREUSED_DISPLAY_COLS,
                        "total_rows": total_rows,
                        "matched_files": len(results),
                    })
                else:
                    if not flash.get("warning"):
                        flash["warning"] = "No matching records found after applying filters."

            except ValueError as e:
                flash["error"] = str(e)
            except Exception as e:
                flash["error"] = f"Error processing files: {e}"

        # Always redirect — eliminates the browser resubmission dialog on every POST path
        request.session["_flash"] = flash
        return redirect("upload")

    # GET — pop flash once so refresh shows a blank page
    context = request.session.pop("_flash", {})
    return render(request, "analysis/upload.html", context)


def clear_results(request):
    request.session.pop("_flash", None)
    request.session.pop("result_data", None)
    return redirect("upload")


def download_excel(request):
    data = request.session.get("result_data")
    if not data:
        return HttpResponse("No result data found. Please run the analysis first.", status=400)

    columns = data["columns"]
    rows = data["rows"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WhereUsed Analysis"

    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill("solid", fgColor="2563EB")
    header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = openpyxl.styles.Side(style="thin", color="D1D5DB")
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 30
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    alt_fill = openpyxl.styles.PatternFill("solid", fgColor="EFF6FF")
    no_fill = openpyxl.styles.PatternFill(fill_type=None)
    for row_idx, row in enumerate(rows, start=2):
        row_fill = alt_fill if row_idx % 2 == 0 else no_fill
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = openpyxl.styles.Alignment(vertical="center")

    for col_idx, col_name in enumerate(columns, start=1):
        col_values = [col_name] + [r[col_idx - 1] for r in rows]
        max_len = max(len(str(v)) for v in col_values)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="whereused_analysis.xlsx"'
    return response
